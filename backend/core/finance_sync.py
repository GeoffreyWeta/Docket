"""Pull the post-award ledger in from the finance system.

The organisation runs Dynamics NAV today and expects to move to Business
Central. That migration is the reason this module has the shape it has: the
import is split into a **field mapping** that knows what one system calls
things, and an **apply** step that knows nothing about any system at all.
Moving to Business Central means writing a second mapping next to the first,
switching a setting, and leaving every row already imported attributable to the
system that supplied it. It does not mean touching the models, the rollups, or
the dashboards.

    NAV export ─┐
                ├─► normalise() ─► NORMAL ROWS ─► apply() ─► the mirror
    BC OData ───┘   (per-source)    (system-free)  (shared)

The normal row is the contract between the two halves. It is a plain dict with
documented keys, base-currency money, and epoch-millisecond dates — the same
vocabulary the rest of DOCKET already speaks.

**What this module refuses to do.** It does not invent a link. A NAV invoice
whose contract reference matches nothing here is imported unlinked and counted
as unlinked, because the finding "412 invoices reference contracts we do not
hold" is the reason Finance asked for this page. Guessing the nearest contract
would replace a visible integration gap with an invisible reporting error.
"""
import csv
import io
import re
from datetime import datetime, timezone

from django.db import transaction

from .models import (Contract, FxRate, GoodsReceipt, Invoice, Payment, PurchaseOrder,
                     SourceSync, Supplier, Tender)
from .util import now_ms, rid

ENTITIES = ("contract", "po", "grn", "invoice", "payment", "fx")

BASE_CCY = "NGN"


# ============================================================ normal row shape
#
# Every adapter emits dicts in these shapes. Keys absent means "the source did
# not say", which is different from zero and is preserved as None.
#
#   contract  external_id ref title supplier_code tender_ref value original_value
#             currency fx_rate signed_at starts_at ends_at status change_orders
#             department cost_centre project region funding_source
#   po        external_id ref description supplier_code contract_ref tender_ref
#             amount currency fx_rate raised_at raised_by approved_at approved_by status
#   grn       external_id ref po_ref amount currency fx_rate received_at received_by note
#   invoice   external_id supplier_ref supplier_code contract_ref po_ref grn_ref
#             amount currency fx_rate invoiced_at received_at due_at approved_at
#             approved_by status hold_reason discount_pct discount_days
#   payment   external_id ref invoice_ref supplier_code amount currency fx_rate
#             paid_at method discount_taken
#   fx        currency at rate


# ==================================================================== adapters

class Adapter:
    """A finance system's vocabulary. Subclasses map its columns onto normal
    rows and nothing else — no database access, no business rules."""

    key = ""
    label = ""

    def normalise(self, entity, payload):
        raise NotImplementedError


def _num(v):
    """A money or rate cell. Absorbs thousands separators, currency marks,
    trailing minus and accounting parentheses; returns None for a blank so a
    missing figure never silently becomes zero."""
    s = str(v if v is not None else "").strip()
    if not s or s.upper() in ("NIL", "N/A", "NA", "-", "NULL", "NONE"):
        return None
    neg = s.startswith("(") and s.endswith(")") or s.endswith("-")
    s = re.sub(r"[^0-9.\-]", "", s.strip("()"))
    if s in ("", "-", "."):
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return -abs(n) if neg else n


def _money(v):
    n = _num(v)
    return None if n is None else int(round(n))


# NAV exports carry whichever of these the person who built the view chose.
_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d %b %Y",
                 "%d-%b-%Y", "%d/%m/%y", "%Y/%m/%d", "%d.%m.%Y")


def _date_ms(v):
    """A date cell as epoch milliseconds, UTC midnight. Returns None rather than
    a guess: an invoice with an unreadable date must not be aged against today."""
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = float(v)
        # Already epoch ms, or epoch seconds, or an Excel serial day number.
        if n > 1e11:
            return int(n)
        if n > 1e9:
            return int(n * 1000)
        if 20000 < n < 80000:
            return int((n - 25569) * 86_400_000)
        return None
    s = str(v).strip()
    if not s:
        return None
    s = s.split("T")[0].split(" ")[0] if re.match(r"^\d{4}-\d{2}-\d{2}", s) else s
    for fmt in _DATE_FORMATS:
        try:
            return int(datetime.strptime(s, fmt).replace(tzinfo=timezone.utc).timestamp() * 1000)
        except ValueError:
            continue
    return None


def _pick(row, *names):
    """First non-empty value among several possible column names, matched
    case- and space-insensitively. NAV views are built by hand and the same
    field appears as "Vendor No.", "VENDOR NO", and "Vendor_No" across three
    exports of the same table."""
    flat = {re.sub(r"[^a-z0-9]", "", str(k).lower()): v for k, v in row.items()}
    for n in names:
        v = flat.get(re.sub(r"[^a-z0-9]", "", n.lower()))
        if v not in (None, ""):
            return v
    return None


class NavAdapter(Adapter):
    """Dynamics NAV, read from a spreadsheet or CSV export of the standard
    tables. Field names follow NAV's own captions, with the common hand-built
    variants accepted alongside them."""

    key = "nav"
    label = "Dynamics NAV"

    def normalise(self, entity, payload):
        fn = getattr(self, f"_{entity}", None)
        if fn is None:
            return []
        return [r for r in (fn(row) for row in payload) if r and r.get("external_id")]

    def _contract(self, r):
        value = _money(_pick(r, "Contract Value", "Amount", "Line Amount"))
        original = _money(_pick(r, "Original Value", "Contract Value")) or value
        return {
            "external_id": str(_pick(r, "No.", "Contract No.", "Document No.") or "").strip(),
            "ref": str(_pick(r, "Contract No.", "No.") or "").strip(),
            "title": str(_pick(r, "Description", "Contract Name") or "").strip(),
            "supplier_code": str(_pick(r, "Vendor No.", "Buy-from Vendor No.", "Nav Code") or "").strip(),
            "tender_ref": str(_pick(r, "External Document No.", "Tender Ref", "Your Reference") or "").strip(),
            "value": value,
            "original_value": original,
            "currency": (str(_pick(r, "Currency Code") or BASE_CCY).strip() or BASE_CCY).upper(),
            "fx_rate": _num(_pick(r, "Currency Factor", "Exchange Rate")),
            "signed_at": _date_ms(_pick(r, "Document Date", "Signed Date", "Posting Date")),
            "starts_at": _date_ms(_pick(r, "Starting Date", "Start Date")),
            "ends_at": _date_ms(_pick(r, "Ending Date", "Expiry Date", "End Date")),
            "status": self._contract_status(_pick(r, "Status")),
            "department": str(_pick(r, "Global Dimension 1 Code", "Department Code", "Department") or "").strip(),
            "cost_centre": str(_pick(r, "Global Dimension 2 Code", "Cost Centre", "Cost Center") or "").strip(),
            "project": str(_pick(r, "Job No.", "Project", "Project Code") or "").strip(),
            "region": str(_pick(r, "Area", "Region", "Location Code") or "").strip(),
            "funding_source": str(_pick(r, "Source Code", "Funding Source", "Budget Code") or "").strip(),
        }

    @staticmethod
    def _contract_status(v):
        s = str(v or "").strip().lower()
        if s.startswith("clos"):
            return "closed"
        if s.startswith("term") or s.startswith("cancel"):
            return "terminated"
        if s.startswith("expir"):
            return "expired"
        return "active"

    def _po(self, r):
        return {
            "external_id": str(_pick(r, "No.", "Order No.", "Document No.") or "").strip(),
            "ref": str(_pick(r, "No.", "Order No.") or "").strip(),
            "description": str(_pick(r, "Description", "Posting Description") or "").strip(),
            "supplier_code": str(_pick(r, "Buy-from Vendor No.", "Vendor No.") or "").strip(),
            "contract_ref": str(_pick(r, "Contract No.", "Blanket Order No.") or "").strip(),
            "tender_ref": str(_pick(r, "External Document No.", "Your Reference") or "").strip(),
            "amount": _money(_pick(r, "Amount Including VAT", "Amount", "Line Amount")),
            "currency": (str(_pick(r, "Currency Code") or BASE_CCY).strip() or BASE_CCY).upper(),
            "fx_rate": _num(_pick(r, "Currency Factor", "Exchange Rate")),
            "raised_at": _date_ms(_pick(r, "Order Date", "Document Date", "Posting Date")),
            "raised_by": str(_pick(r, "Purchaser Code", "Assigned User ID", "Raised By") or "").strip(),
            "approved_at": _date_ms(_pick(r, "Approval Date", "Approved Date")),
            "approved_by": str(_pick(r, "Approver ID", "Approved By") or "").strip(),
            "status": str(_pick(r, "Status") or "open").strip().lower(),
        }

    def _grn(self, r):
        return {
            "external_id": str(_pick(r, "No.", "Receipt No.", "Document No.") or "").strip(),
            "ref": str(_pick(r, "No.", "Receipt No.") or "").strip(),
            "po_ref": str(_pick(r, "Order No.", "Purchase Order No.") or "").strip(),
            "amount": _money(_pick(r, "Amount", "Line Amount")),
            "currency": (str(_pick(r, "Currency Code") or BASE_CCY).strip() or BASE_CCY).upper(),
            "fx_rate": _num(_pick(r, "Currency Factor", "Exchange Rate")),
            "received_at": _date_ms(_pick(r, "Posting Date", "Receipt Date", "Document Date")),
            "received_by": str(_pick(r, "Assigned User ID", "Received By") or "").strip(),
            "note": str(_pick(r, "Description") or "").strip(),
        }

    def _invoice(self, r):
        return {
            "external_id": str(_pick(r, "No.", "Document No.", "Entry No.") or "").strip(),
            "supplier_ref": str(_pick(r, "Vendor Invoice No.", "External Document No.") or "").strip(),
            "supplier_code": str(_pick(r, "Buy-from Vendor No.", "Vendor No.", "Pay-to Vendor No.") or "").strip(),
            "contract_ref": str(_pick(r, "Contract No.") or "").strip(),
            "po_ref": str(_pick(r, "Order No.", "Purchase Order No.") or "").strip(),
            "grn_ref": str(_pick(r, "Receipt No.", "Package Tracking No.") or "").strip(),
            "amount": _money(_pick(r, "Amount Including VAT", "Amount")),
            "currency": (str(_pick(r, "Currency Code") or BASE_CCY).strip() or BASE_CCY).upper(),
            "fx_rate": _num(_pick(r, "Currency Factor", "Exchange Rate")),
            "invoiced_at": _date_ms(_pick(r, "Document Date", "Invoice Date")),
            "received_at": _date_ms(_pick(r, "Posting Date", "Received Date", "Document Date")),
            "due_at": _date_ms(_pick(r, "Due Date")),
            "approved_at": _date_ms(_pick(r, "Approval Date", "Approved Date")),
            "approved_by": str(_pick(r, "Approver ID", "Approved By") or "").strip(),
            "status": str(_pick(r, "Status", "Document Status") or "received").strip().lower(),
            "hold_reason": str(_pick(r, "On Hold", "Hold Reason") or "").strip(),
            "discount_pct": _num(_pick(r, "Payment Discount %", "Pmt. Discount %")) or 0,
            "discount_days": int(_num(_pick(r, "Discount Days")) or 0),
        }

    def _payment(self, r):
        return {
            "external_id": str(_pick(r, "No.", "Document No.", "Entry No.") or "").strip(),
            "ref": str(_pick(r, "No.", "Document No.", "Cheque No.") or "").strip(),
            "invoice_ref": str(_pick(r, "Applies-to Doc. No.", "Invoice No.", "Applies-to ID") or "").strip(),
            "supplier_code": str(_pick(r, "Vendor No.", "Pay-to Vendor No.") or "").strip(),
            "amount": _money(_pick(r, "Amount", "Debit Amount")),
            "currency": (str(_pick(r, "Currency Code") or BASE_CCY).strip() or BASE_CCY).upper(),
            "fx_rate": _num(_pick(r, "Currency Factor", "Exchange Rate")),
            "paid_at": _date_ms(_pick(r, "Posting Date", "Payment Date", "Document Date")),
            "method": str(_pick(r, "Payment Method Code", "Bal. Account No.") or "").strip(),
            "discount_taken": _money(_pick(r, "Pmt. Disc. Rcd.(LCY)", "Payment Discount")) or 0,
        }

    def _fx(self, r):
        return {
            "external_id": f"{_pick(r, 'Currency Code')}:{_pick(r, 'Starting Date')}",
            "currency": (str(_pick(r, "Currency Code") or "").strip()).upper(),
            "at": _date_ms(_pick(r, "Starting Date", "Date")),
            "rate": _num(_pick(r, "Relational Exch. Rate Amount", "Exchange Rate Amount", "Rate")),
        }


class BusinessCentralAdapter(NavAdapter):
    """Business Central, read from the standard API v2.0 payload.

    Subclasses NAV rather than repeating it: BC *is* NAV's data model with the
    captions renamed to camelCase and dates already ISO-8601. `_pick` strips
    punctuation and case before matching, so "Document Date" already finds
    `documentDate` — which means the inherited mappings work unchanged and only
    the genuine renames need listing below. When the migration happens, this
    class is the diff.

    Transport is deliberately not here. `normalise` takes rows that have already
    been fetched, so the OAuth client-credentials call and the OData paging live
    at the edge (`fetch_bc`, below) and this class stays testable against a
    saved payload.
    """

    key = "bc"
    label = "Business Central"

    # BC renames a handful of fields outright. Everything else falls through to
    # the NAV mapping by punctuation-insensitive match.
    RENAMES = {
        "Vendor No.": "vendorNumber",
        "Buy-from Vendor No.": "vendorNumber",
        "Pay-to Vendor No.": "payToVendorNumber",
        "Vendor Invoice No.": "vendorInvoiceNumber",
        "External Document No.": "externalDocumentNumber",
        "Amount Including VAT": "totalAmountIncludingTax",
        "Amount": "totalAmountExcludingTax",
        "Currency Code": "currencyCode",
        "Currency Factor": "currencyFactor",
        "Due Date": "dueDate",
        "Document Date": "documentDate",
        "Posting Date": "postingDate",
        "No.": "number",
        "Order No.": "orderNumber",
        "Status": "status",
    }

    def normalise(self, entity, payload):
        # Re-key BC's names to the NAV captions the inherited mappings expect,
        # keeping the original keys too so a field only BC has is still reachable.
        alias = {v.lower(): k for k, v in self.RENAMES.items()}
        rows = []
        for row in payload:
            merged = dict(row)
            for k, v in row.items():
                nav_name = alias.get(str(k).lower())
                if nav_name and merged.get(nav_name) in (None, ""):
                    merged[nav_name] = v
            rows.append(merged)
        return super().normalise(entity, rows)


ADAPTERS = {a.key: a() for a in (NavAdapter, BusinessCentralAdapter)}


def adapter_for(source):
    return ADAPTERS.get(source, ADAPTERS["nav"])


def fetch_bc(entity, *, tenant, company, token, base="https://api.businesscentral.dynamics.com"):
    """Fetch one entity from Business Central's API v2.0, following @odata.nextLink.

    Separated from the adapter so nothing about the mapping depends on having
    credentials: the migration can be rehearsed end to end against a saved
    payload, and this function is the only part that needs a tenant to test.
    `token` is a bearer token obtained by the caller through the normal
    client-credentials flow; deliberately not minted here, so no secret has to
    be stored in this codebase.
    """
    import json
    import urllib.request

    path = {"contract": "purchaseOrders", "po": "purchaseOrders", "grn": "purchaseReceipts",
            "invoice": "purchaseInvoices", "payment": "vendorPayments",
            "fx": "currencyExchangeRates"}.get(entity)
    if not path:
        return []
    url = f"{base}/v2.0/{tenant}/production/api/v2.0/companies({company})/{path}"
    out = []
    while url:
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}",
                                                   "Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=60) as fh:
            body = json.loads(fh.read().decode())
        out.extend(body.get("value") or [])
        url = body.get("@odata.nextLink")
    return out


# ===================================================================== reading

def rows_from_csv(text):
    return list(csv.DictReader(io.StringIO(text)))


def rows_from_upload(fh):
    """A CSV or XLSX upload as a list of dicts. XLSX goes through openpyxl if it
    is installed, which it is for the vendor register import."""
    name = (getattr(fh, "name", "") or "").lower()
    data = fh.read()
    if name.endswith(".csv") or not name.endswith((".xlsx", ".xlsm")):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return rows_from_csv(data.decode(enc))
            except UnicodeDecodeError:
                continue
        return []
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c or "").strip() for c in next(rows, [])]
    return [dict(zip(header, r)) for r in rows if any(c not in (None, "") for c in r)]


# ======================================================================= apply

def _fx_rate(row, rates):
    """The rate to convert this row at: the one the source recorded, or today's
    observation for that currency, or 1. Recorded beats current — a contract
    struck at last year's rate is a commitment at last year's rate, and
    revaluing it silently would erase the exposure this page exists to show."""
    ccy = (row.get("currency") or BASE_CCY).upper()
    if ccy == BASE_CCY:
        return 1.0
    given = row.get("fx_rate")
    if given and given > 0:
        # NAV's "Currency Factor" is the reciprocal of a rate when it is < 1.
        return given if given >= 1 else (1 / given)
    return rates.get(ccy) or 1.0


def _amounts(row, rates):
    """(base, source, currency, rate) for one normal row."""
    src = row.get("amount") if "amount" in row else row.get("value")
    src = int(src or 0)
    ccy = (row.get("currency") or BASE_CCY).upper()
    rate = _fx_rate(row, rates)
    return int(round(src * rate)), src, ccy, rate


def _supplier_index():
    """{nav code / previous code / normalised name: Supplier}.

    The register already carries each vendor's NAV code, and the importer kept
    any code a vendor was previously registered under — so a ledger row keyed to
    the old code still lands on the right vendor instead of creating a second.
    """
    idx = {}
    for s in Supplier.objects.all():
        for code in filter(None, [s.code, (s.registry or {}).get("previousCode")]):
            idx[str(code).strip().upper()] = s
        idx.setdefault(re.sub(r"[^A-Z0-9]", "", (s.name or "").upper()), s)
    return idx


def _tender_index():
    return {t.ref.strip().upper(): t for t in Tender.objects.all() if t.ref}


def apply_rows(source, entity, rows, *, now=None):
    """Upsert normal rows into the mirror. Returns a report, and never raises on
    one bad row: an import that abandons 4,000 good rows because row 91 has a
    date nobody can parse is an import that never runs."""
    now = now or now_ms()
    rates = FxRate.latest()
    report = {"seen": len(rows), "written": 0, "skipped": [], "unlinked": 0}
    fn = {"contract": _apply_contract, "po": _apply_po, "grn": _apply_grn,
          "invoice": _apply_invoice, "payment": _apply_payment, "fx": _apply_fx}.get(entity)
    if fn is None:
        report["skipped"].append({"row": None, "why": f"unknown entity {entity!r}"})
        return report

    ctx = {"suppliers": _supplier_index(), "tenders": _tender_index(),
           "rates": rates, "now": now, "source": source}
    for row in rows:
        try:
            with transaction.atomic():
                linked = fn(row, ctx)
            report["written"] += 1
            if linked is False:
                report["unlinked"] += 1
        except Exception as exc:                                  # noqa: BLE001
            report["skipped"].append({"row": row.get("external_id"), "why": str(exc)[:200]})
    return report


def _base(row, ctx, model):
    """The provenance fields every mirrored row carries, plus its stable key."""
    ext = str(row.get("external_id") or "").strip()
    existing = model.objects.filter(source=ctx["source"], external_id=ext).first()
    return existing, {"source": ctx["source"], "external_id": ext, "synced_at": ctx["now"]}


def _supplier(row, ctx):
    code = str(row.get("supplier_code") or "").strip().upper()
    return ctx["suppliers"].get(code)


def _apply_contract(row, ctx):
    existing, prov = _base(row, ctx, Contract)
    amount, src, ccy, rate = _amounts({**row, "amount": row.get("value")}, ctx["rates"])
    original = row.get("original_value")
    tender = ctx["tenders"].get(str(row.get("tender_ref") or "").strip().upper())
    fields = {
        **prov,
        "ref": row.get("ref") or prov["external_id"],
        "title": row.get("title") or "",
        "supplier": _supplier(row, ctx),
        "tender": tender,
        "amount": amount, "amount_src": src, "currency": ccy, "fx_rate": rate,
        "original_value": int(round((original if original is not None else src) * rate)),
        "signed_at": row.get("signed_at"),
        "starts_at": row.get("starts_at"),
        "ends_at": row.get("ends_at"),
        "status": row.get("status") or "active",
        "department": row.get("department") or "",
        "cost_centre": row.get("cost_centre") or "",
        "project": row.get("project") or "",
        "region": row.get("region") or "",
        "funding_source": row.get("funding_source") or "",
    }
    # Dimensions the ledger did not carry fall back to the tender's, which is
    # the same commitment described by the side of the house that recorded them.
    if tender:
        for k in ("department", "cost_centre", "project", "region", "funding_source"):
            if not fields[k]:
                fields[k] = getattr(tender, k, "") or ""
    if existing:
        for k, v in fields.items():
            setattr(existing, k, v)
        existing.save()
    else:
        Contract.objects.create(id=rid("ct"), **fields)
    return tender is not None


def _apply_po(row, ctx):
    existing, prov = _base(row, ctx, PurchaseOrder)
    amount, src, ccy, rate = _amounts(row, ctx["rates"])
    contract = Contract.objects.filter(ref=str(row.get("contract_ref") or "").strip()).first() \
        if row.get("contract_ref") else None
    fields = {
        **prov,
        "ref": row.get("ref") or prov["external_id"],
        "description": row.get("description") or "",
        "supplier": _supplier(row, ctx),
        "contract": contract,
        "tender": ctx["tenders"].get(str(row.get("tender_ref") or "").strip().upper()),
        "amount": amount, "amount_src": src, "currency": ccy, "fx_rate": rate,
        "raised_at": row.get("raised_at"),
        "raised_by": row.get("raised_by") or "",
        "approved_at": row.get("approved_at"),
        "approved_by": row.get("approved_by") or "",
        "status": row.get("status") or "open",
    }
    if existing:
        for k, v in fields.items():
            setattr(existing, k, v)
        existing.save()
    else:
        PurchaseOrder.objects.create(id=rid("po"), **fields)
    return contract is not None


def _apply_grn(row, ctx):
    existing, prov = _base(row, ctx, GoodsReceipt)
    amount, src, ccy, rate = _amounts(row, ctx["rates"])
    order = PurchaseOrder.objects.filter(ref=str(row.get("po_ref") or "").strip()).first() \
        if row.get("po_ref") else None
    fields = {
        **prov,
        "ref": row.get("ref") or prov["external_id"],
        "order": order,
        "amount": amount, "amount_src": src, "currency": ccy, "fx_rate": rate,
        "received_at": row.get("received_at"),
        "received_by": row.get("received_by") or "",
        "note": row.get("note") or "",
    }
    if existing:
        for k, v in fields.items():
            setattr(existing, k, v)
        existing.save()
    else:
        GoodsReceipt.objects.create(id=rid("gr"), **fields)
    return order is not None


def _apply_invoice(row, ctx):
    existing, prov = _base(row, ctx, Invoice)
    amount, src, ccy, rate = _amounts(row, ctx["rates"])
    order = PurchaseOrder.objects.filter(ref=str(row.get("po_ref") or "").strip()).first() \
        if row.get("po_ref") else None
    contract = Contract.objects.filter(ref=str(row.get("contract_ref") or "").strip()).first() \
        if row.get("contract_ref") else None
    if contract is None and order is not None:
        contract = order.contract
    receipt = GoodsReceipt.objects.filter(ref=str(row.get("grn_ref") or "").strip()).first() \
        if row.get("grn_ref") else None
    if receipt is None and order is not None:
        receipt = order.receipts.first()
    fields = {
        **prov,
        "supplier_ref": row.get("supplier_ref") or "",
        "supplier": _supplier(row, ctx),
        "contract": contract, "order": order, "receipt": receipt,
        "amount": amount, "amount_src": src, "currency": ccy, "fx_rate": rate,
        "invoiced_at": row.get("invoiced_at"),
        "received_at": row.get("received_at") or row.get("invoiced_at"),
        "due_at": row.get("due_at"),
        "approved_at": row.get("approved_at"),
        "approved_by": row.get("approved_by") or "",
        "status": _invoice_status(row.get("status")),
        "hold_reason": row.get("hold_reason") or "",
        "discount_pct": row.get("discount_pct") or 0,
        "discount_days": row.get("discount_days") or 0,
    }
    if existing:
        for k, v in fields.items():
            setattr(existing, k, v)
        existing.save()
    else:
        Invoice.objects.create(id=rid("iv"), **fields)
    return contract is not None


def _invoice_status(v):
    s = str(v or "").strip().lower()
    if "paid" in s and "part" in s:
        return "part_paid"
    if "paid" in s or "closed" in s:
        return "paid"
    if "reject" in s:
        return "rejected"
    if "approv" in s or "released" in s or "open" in s:
        return "approved"
    return "received"


def _apply_payment(row, ctx):
    existing, prov = _base(row, ctx, Payment)
    amount, src, ccy, rate = _amounts(row, ctx["rates"])
    ref = str(row.get("invoice_ref") or "").strip()
    invoice = None
    if ref:
        invoice = (Invoice.objects.filter(source=ctx["source"], external_id=ref).first()
                   or Invoice.objects.filter(supplier_ref=ref).first())
    fields = {
        **prov,
        "ref": row.get("ref") or prov["external_id"],
        "invoice": invoice,
        "supplier": (invoice.supplier if invoice else None) or _supplier(row, ctx),
        # NAV posts vendor payments as negative amounts; the mirror stores what
        # left as a positive figure so nothing downstream has to remember a sign.
        "amount": abs(amount), "amount_src": abs(src), "currency": ccy, "fx_rate": rate,
        "paid_at": row.get("paid_at"),
        "method": row.get("method") or "",
        "discount_taken": abs(row.get("discount_taken") or 0),
    }
    if existing:
        for k, v in fields.items():
            setattr(existing, k, v)
        existing.save()
    else:
        Payment.objects.create(id=rid("pm"), **fields)
    return invoice is not None


def _apply_fx(row, ctx):
    ccy, at, rate = row.get("currency"), row.get("at"), row.get("rate")
    if not ccy or not at or not rate:
        raise ValueError("incomplete rate row")
    FxRate.objects.update_or_create(currency=ccy.upper(), at=at,
                                    defaults={"rate": rate, "source": ctx["source"]})
    return True


# ================================================================ orchestration

def sync(source, entity, payload, *, now=None):
    """Normalise a payload through `source`'s adapter and apply it.

    The SourceSync row is written whatever happens, success or failure, because
    "the last attempt failed" is the fact the dashboard needs in order to stop
    presenting a three-week-old ledger as current.
    """
    now = now or now_ms()
    state, _ = SourceSync.objects.get_or_create(source=source, entity=entity)
    state.last_attempt = now
    try:
        rows = adapter_for(source).normalise(entity, payload)
        report = apply_rows(source, entity, rows, now=now)
        state.rows_seen = report["seen"]
        state.rows_written = report["written"]
        state.error = ""
        if report["skipped"]:
            state.error = f"{len(report['skipped'])} row(s) skipped: {report['skipped'][0]['why']}"
        state.last_success = now
        state.save()
        return report
    except Exception as exc:                                      # noqa: BLE001
        state.error = str(exc)[:300]
        state.save()
        raise


def sync_state():
    """What every feed last did, for the staleness banner."""
    return [{"source": s.source, "entity": s.entity, "lastAttempt": s.last_attempt,
             "lastSuccess": s.last_success, "rowsSeen": s.rows_seen,
             "rowsWritten": s.rows_written, "error": s.error}
            for s in SourceSync.objects.all().order_by("source", "entity")]


def freshness():
    """(oldest successful sync across the feeds that have run, feeds never run).

    The oldest, not the newest: a page is only as current as its stalest feed,
    and reporting the newest would let one healthy feed vouch for five dead ones.
    """
    rows = list(SourceSync.objects.all())
    ran = [r for r in rows if r.last_success]
    never = [r.entity for r in rows if not r.last_success]
    missing = [e for e in ENTITIES if e not in {r.entity for r in rows}]
    return (min((r.last_success for r in ran), default=None), sorted(set(never + missing)))
